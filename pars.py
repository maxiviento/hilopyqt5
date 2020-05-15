from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QObject, QThread, pyqtSignal, pyqtSlot

from bs4 import BeautifulSoup
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException
import openpyxl as excel
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import time
import os
import json
import numpy as np

class Parseador():
    def __init__(self, nav, signal):

      self.countChanged = signal

      self.nav = nav
      self.driver = self.nav.abrir_Cr()
      self.username = os.environ['username']
      self.directorio = r"C:\Users\{}\Documents".format(self.username)
      self.fecha_hora = self.fecha_y_hora()
      self.hoy = datetime.now().date()
      #definimos donde guardo datos del parseo, hoy es excel mañana bbdd
      self.ruta_parseo = r"{}\parseos\peliculas_{}.xlsx".format(self.directorio, self.fecha_hora)
      self.excel_parseo = self.crea_archivo_parseo()
      self.val = 0


    def fecha_y_hora (self):
      hoy = str(datetime.now().date())
      ahora = str(datetime.now().time())
      ahora = ahora.replace(":","")
      ahora = ahora[:6]
      fecha_hora = hoy + ahora
      return fecha_hora
    
    def crea_archivo_parseo (self):
      try:
        os.stat(self.directorio + "\parseos")
      except:
        os.mkdir(self.directorio + "\parseos")

      wb=Workbook()
      wb.save(self.ruta_parseo)
      self.excel_parseo = load_workbook(self.ruta_parseo)
      self.parseo = self.excel_parseo.active
      self.parseo['A1'] = 'titleColumn'
      self.parseo['B1'] = 'peli'
      self.parseo['C1'] = 'Año'

      self.excel_parseo.save(self.ruta_parseo)
      return self.excel_parseo

  
    def guarda_archivo_parseo(self):
      self.excel_parseo.save(self.ruta_parseo)

    def emitidos(self):
      return self.val

    def formato_texto(self, texto, index):
      aux = texto[::-1]
      busca_par_ab = aux.find('(')
      año = aux[1:busca_par_ab]
      año = año[::-1]
      año = año[:4]
      aux = aux[busca_par_ab + 2:]
      busca_salto_linea = aux.find("\n")
      titulo = aux[:busca_salto_linea - 6]
      titulo = titulo[::-1]
      index = str(index + 1)
      return index, titulo, año

    
    @pyqtSlot()
    def parsea (self):
      soup = BeautifulSoup(self.driver.page_source, 'html.parser')
      lista_titulos = soup.findAll('td', {'class' : 'titleColumn'})
      total = len(lista_titulos)
      for index, i in enumerate(lista_titulos, start=1):
        self.val = 100 * float(index)/float(total)
        self.countChanged.emit(self.val)
        print(str(self.val), i.text)
        tupla_p_excel = self.formato_texto(i.text, index)
        self.parseo.append(tupla_p_excel)
      
      self.guarda_archivo_parseo()

            